"""
Skoda DSR Generator
Extracts shipment data from BL PDFs and Invoices, generating DSR Excel (container-wise).

Features:
- Maersk & Hapag-Lloyd BL formats.
- Invoices extracted from PDF filenames.
- tkcalendar for date selection.
- GUI layout matches Nagarkot VW/Audi tool standards.
"""

from __future__ import annotations

import re
import logging
import os
import sys
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Optional

try:
    from tkcalendar import Calendar
except ImportError:
    # This will still show red if the IDE doesn't see the package, 
    # but the runtime check is here.
    Calendar = None 

import fitz  # PyMuPDF
import openpyxl
import openpyxl.utils
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageTk

from bl_parser import ContainerRecord, parse_bl
from zoho_api import ZohoCreatorAPI


# ─── PyInstaller Resource Helper ─────────────────────────────────────────────

def resource_path(relative_path: str) -> Path:
    """Get absolute path to resource, works for dev and for PyInstaller."""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = Path(sys._MEIPASS)  # type: ignore[attr-defined]
    except AttributeError:
        base_path = Path(os.path.abspath("."))
    return base_path / relative_path


# ─── Constants ───────────────────────────────────────────────────────────────

# For bundled resources (logo) — looks inside the PyInstaller temp dir
LOGO_PATH = resource_path("Nagarkot Logo.png")

# For files that live next to the exe (master DSR, .env, output files)
# When frozen, sys.executable points to the exe; otherwise use __file__
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = Path(sys.executable).resolve().parent
else:
    SCRIPT_DIR = Path(__file__).resolve().parent

MASTER_FILE_PATH = SCRIPT_DIR / "SKODA_MASTER_DSR.xlsx"

# DSR column headers in order (A → BF = 58 columns)
DSR_HEADERS: list[str] = [
    "User",                           # A
    "Pre-alert Receive date",         # B
    "Month",                          # C
    "FF/ Shipping Line",              # D
    "Port of Loading",                # E
    "Vessel Name",                    # F
    "BL Date",                        # G
    "Vessel ETA",                     # H
    "CHA Job No.",                    # I
    "Container No.",                  # J
    "Size (20'40' LCL)",              # K
    "Container Type (HQ,DV,SD)",      # L
    "Current Status",                 # M
    "BL No.",                         # N
    "Supplier Name",                  # O
    "Invoice No.",                    # P
    "INCO",                           # Q
    "No.of Pkg.",                     # R
    "GrossWt",                        # S
    "CFS Name",                       # T
    "IGM No.",                        # U
    "IGM No. Date",                   # V
    "IGM Inward Date",                # W
    "B/E No",                         # X
    "B/E Date",                       # Y
    "AO Ass",                         # Z
    "AC Assess",                      # AA
    "RMS/ Examine",                   # AB
    "Duty Request recd from CHA",     # AC
    "Duty Paid date",                 # AD
    "Assessable Value",               # AE
    "Debit Duty (RODTEP)",            # AF
    "Total Duty",                     # AG
    "DUTY%",                          # AH
    "STAMP DUTY",                     # AI
    "Interest (IfAny)",               # AJ
    "Penalty (Ifany)",                # AK
    "Reason for Interest/Penalty",    # AL
    "OOC Date",                       # AM
    "Dispatch date to plant/WH",      # AN
    "Remarks (Daywise Cronology)",    # AO
    "Clearnace TAT",                  # AP
    "Reason for Clearance TAT delay", # AQ
    "E-Waybill No.",                  # AR
    "Detention/Demurrage (IfAny)",    # AS
    "Total BCD Value",                # AT
    "Total SWS Value",                # AU
    "Total IGST Value",               # AV
    "CHA JOB NO",                     # AW
    "Transporter",                    # AX
    "STAMP DUTY PAID DT",             # AY
    "Under Protest",                  # AZ
    "BOE filing TAT",                 # BA
    "Reason for BOE filing TAT delay",# BB
    "Conatainer arrival date in CFS", # BC
    "OOC COPY RECD YES/NO",           # BD
    "Remarks",                        # BE
    "SIMS Registration date",         # BF
]

MONTH_MAP = {
    "01": "JAN", "02": "FEB", "03": "MAR", "04": "APR",
    "05": "MAY", "06": "JUN", "07": "JUL", "08": "AUG",
    "09": "SEP", "10": "OCT", "11": "NOV", "12": "DEC",
}

# ─── Theme ───────────────────────────────────────────────────────────────────

NAGARKOT_BLUE = "#1B3A5C"
BTN_BLUE = "#0056b3"
WHITE = "#FFFFFF"
BG_LIGHT = "#F8F9FA"

class BetterDateEntry(tk.Frame):
    """Custom DatePicker without the Windows scaling double-calendar bug in DateEntry."""
    def __init__(self, master, width=15, bg="white", **kwargs):
        super().__init__(master, bg=bg)
        self.entry_var = tk.StringVar()
        self.entry = ttk.Entry(self, textvariable=self.entry_var, width=width, font=("Segoe UI", 9))
        self.entry.pack(side="left")
        
        self.btn = tk.Button(self, text="📅", command=self._popup, relief="flat", bg="white", cursor="hand2")
        self.btn.pack(side="left", padx=2)
        
    def _popup(self):
        top = tk.Toplevel(self)
        top.overrideredirect(True)
        top.attributes("-topmost", True)
        
        x = self.entry.winfo_rootx()
        y = self.entry.winfo_rooty() + self.entry.winfo_height() + 2
        top.geometry(f"+{x}+{y}")
        
        frame = tk.Frame(top, highlightbackground="#0056b3", highlightthickness=2)
        frame.pack()
        
        if not Calendar:
            messagebox.showerror("Missing Dependency", "Please run: pip install tkcalendar")
            top.destroy()
            return

        cal = Calendar(frame, selectmode="day", date_pattern="yyyy-mm-dd", showweeknumbers=False, 
                       selectmonth=True, selectyear=True,
                       font=("Segoe UI", 9), background="#0056b3", foreground="white",
                       headersbackground="#F8F9FA", headersforeground="black")
        cal.pack()
        
        def on_select(e):
            self.entry_var.set(cal.get_date())
            top.destroy()
            
        cal.bind("<<CalendarSelected>>", on_select)
        def on_scroll(event):
            # Windows/MacOS: event.delta
            if event.delta > 0:
                cal._prev_month()
            else:
                cal._next_month()
        
        cal.bind("<MouseWheel>", on_scroll)
        top.bind("<Escape>", lambda e: top.destroy())
        
        def check_focus():
            if not top.winfo_exists():
                return
            # If focus is None, it might be a momentary transition or a dropdown
            if top.focus_get() is None:
                top.after(200, perform_destroy_if_lost)
            else:
                perform_destroy_if_lost()
        
        def perform_destroy_if_lost():
            try:
                if not top.winfo_exists():
                    return
                new_focus = top.focus_get()
                if new_focus is None:
                    top.destroy()
                    return

                # Check if the new focus widget is a descendant of 'top'
                parent = new_focus
                is_child = False
                while parent:
                    if parent == top:
                        is_child = True
                        break
                    parent = getattr(parent, 'master', None)
                if not is_child:
                    top.destroy()
            except Exception:
                # Catch potential naming/resolution errors from internal tk widgets (like popdown)
                if top.winfo_exists():
                    top.destroy()

        top.bind("<FocusOut>", lambda e: check_focus())
        cal.focus_set()

    def get(self):
        return self.entry_var.get()
        
    def delete(self, first, last=None):
        self.entry.delete(first, last)

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hidetip()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(500, self.showtip)

    def unschedule(self):
        id_ = self.id
        self.id = None
        if id_:
            self.widget.after_cancel(id_)

    def showtip(self, event=None):
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                      background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                      font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class DataPreviewWindow(tk.Toplevel):
    def __init__(self, parent, records: list[ContainerRecord], on_confirm, defaults: dict = None):
        super().__init__(parent)
        self.title("Review & Edit Extracted Data")
        self.geometry("1400x650")
        self.configure(bg=WHITE)
        self.records = records
        self.on_confirm = on_confirm
        self.defaults = defaults or {}
        
        self.transient(parent)
        self.grab_set()
        
        header = tk.Frame(self, bg=WHITE, height=60)
        header.pack(fill="x", side="top")
        tk.Label(header, text="Review & Edit Data", font=("Segoe UI", 16, "bold"), bg=WHITE, fg=BTN_BLUE).pack(pady=5)
        tk.Label(header, text="Double-click any cell to edit details per row.", font=("Segoe UI", 10), bg=WHITE, fg="#6c757d").pack(pady=2)

        frame = tk.Frame(self, bg=WHITE)
        frame.pack(fill="both", expand=True, padx=20, pady=5)
        
        self.col_map = [
            ("user", "User", 90),
            ("user_month", "Month", 60),
            ("pre_alert_date", "Pre-Alert", 90),
            ("vessel_eta", "Vessel ETA", 90),
            ("bl_mode", "Mode", 80),
            ("bl_no", "BL No", 100),
            ("container_no", "Container No", 100),
            ("invoice_nos", "Invoice Nos", 180),
            ("supplier_name", "Supplier", 180),
            ("inco_terms", "INCO", 70),
            ("num_packages", "Packages", 80),
            ("gross_weight", "Gross Wt", 80),
            ("container_size", "Size", 50),
            ("container_type", "Type", 50),
            ("vessel_name", "Vessel Name", 120),
            ("port_of_loading", "POL", 100),
            ("shipping_line", "Line", 100),
            ("bl_date", "BL Date", 90)
        ]
        
        cols = [c[1] for c in self.col_map]
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse")
        
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)
        
        for attr, heading, w in self.col_map:
            self.tree.heading(heading, text=heading)
            self.tree.column(heading, width=w, stretch=False, anchor="w")
            
        for rec in self.records:
            vals = [getattr(rec, attr) for attr, _, _ in self.col_map]
            self.tree.insert("", "end", values=vals)
            
        self.tree.bind("<Double-1>", self._on_double_click)
        
        footer = tk.Frame(self, bg=WHITE, height=50)
        footer.pack(fill="x", side="bottom", padx=20, pady=10)
        
        tk.Button(
            footer, text="Cancel", font=("Segoe UI", 10), width=15, 
            command=self.destroy
        ).pack(side="left")

        tk.Button(
            footer, text="+ Add Row", font=("Segoe UI", 10, "bold"),
            bg="#28A745", fg=WHITE, width=15, cursor="hand2",
            command=self._add_row
        ).pack(side="left", padx=10)

        tk.Button(
            footer, text="✖ Remove Row", font=("Segoe UI", 10),
            bg="#dc3545", fg=WHITE, width=15, cursor="hand2",
            command=self._remove_row
        ).pack(side="left", padx=10)
        
        tk.Button(
            footer, text="Confirm & Submit", font=("Segoe UI", 10, "bold"),
            bg=BTN_BLUE, fg=WHITE, width=20, cursor="hand2",
            command=self._do_confirm
        ).pack(side="right")
        
    def _add_row(self):
        from bl_parser import ContainerRecord
        new_rec = ContainerRecord()
        
        # Priority 1: Existing records for context
        # Priority 2: Passed defaults from main window
        if self.records:
            ref = self.records[0]
            new_rec.user = ref.user
            new_rec.user_month = ref.user_month
            new_rec.pre_alert_date = ref.pre_alert_date
            new_rec.vessel_eta = ref.vessel_eta
            new_rec.bl_mode = ref.bl_mode
        elif self.defaults:
            new_rec.user = self.defaults.get("user", "")
            new_rec.user_month = self.defaults.get("user_month", "")
            new_rec.pre_alert_date = self.defaults.get("pre_alert_date", "")
            new_rec.vessel_eta = self.defaults.get("vessel_eta", "")
            new_rec.bl_mode = self.defaults.get("bl_mode", "")
        
        self.records.append(new_rec)
        vals = [getattr(new_rec, attr) for attr, _, _ in self.col_map]
        item_id = self.tree.insert("", "end", values=vals)
        self.tree.see(item_id)
        self.tree.selection_set(item_id)

    def _remove_row(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Select Row", "Please select a row to remove.")
            return
            
        if messagebox.askyesno("Remove", "Remove the selected row?"):
            # Remove from back to keep indices stable if multiple selected
            items = list(selected)
            items.sort(key=lambda x: self.tree.index(x), reverse=True)
            for item in items:
                idx = self.tree.index(item)
                del self.records[idx]
                self.tree.delete(item)

    def _on_double_click(self, event):
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            return
            
        x, y, width, height = self.tree.bbox(row_id, col_id)
        col_idx = int(col_id[1:]) - 1
        
        value = self.tree.item(row_id, "values")[col_idx]
        
        entry = ttk.Entry(self.tree, font=("Segoe UI", 9))
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, value)
        entry.focus()
        
        def commit(e=None):
            new_val = entry.get()
            vals = list(self.tree.item(row_id, "values"))
            vals[col_idx] = new_val
            self.tree.item(row_id, values=vals)
            
            rec_idx = self.tree.index(row_id)
            attr_name = self.col_map[col_idx][0]
            setattr(self.records[rec_idx], attr_name, new_val)
            
            # If the user edits BL No manually, we must propagate it so Zoho and Excel use the updated values
            if attr_name == "bl_no":
                 old_raw = self.records[rec_idx].raw_mbl_no or ""
                 parts = new_val.split("/")
                 mbl_part = parts[0].strip() if len(parts) > 0 else ""
                 
                 # Logic for raw_mbl_no: if original was Maersk, ensure prefix prefix is attached
                 if old_raw.upper().startswith("MAEU") and not mbl_part.upper().startswith("MAEU"):
                     self.records[rec_idx].raw_mbl_no = f"MAEU{mbl_part}"
                 else:
                     self.records[rec_idx].raw_mbl_no = mbl_part

                 # Also update secondary storage fields so Excel logic works correctly
                 self.records[rec_idx].mbl_no = mbl_part 
                 self.records[rec_idx].hbl_no = parts[1].strip() if len(parts) > 1 else ""
                 # Refresh bl_no again in case prefix logic changed
                 self.records[rec_idx].bl_no = new_val
                 
            entry.destroy()
            
        entry.bind("<Return>", commit)
        entry.bind("<FocusOut>", commit)
        
    def _do_confirm(self):
        self.destroy()
        self.on_confirm()

class DSRGeneratorApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Skoda (SAVWILP) DSR Generator — Nagarkot")
        self.root.state("zoomed")
        self.root.configure(bg=WHITE)
        self.root.minsize(1000, 700)

        # Style configurations
        style = ttk.Style()
        style.theme_use("clam")
        
        # LabelFrame Style
        style.configure("TLabelFrame", background=WHITE)
        style.configure("TLabelFrame.Label", font=("Segoe UI", 10, "bold"), foreground=BTN_BLUE, background=WHITE)
        
        # Treeview Style
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), background="#E9ECEF", foreground="#495057")
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=25)
        
        # State / UI Variables
        self.files_by_dir: dict[Path, list[Path]] = {}
        self.parsed_records: list[ContainerRecord] = []
        self.confirmed_records: list[ContainerRecord] = []
        self.master_dsr_path: Path = MASTER_FILE_PATH
        self.logo_img = None
        
        # Trio UI Variables
        self.var_trio_inv = tk.StringVar()
        self.var_trio_hbl = tk.StringVar()
        self.var_trio_mbl = tk.StringVar()
        self.trio_inv_paths: list[str] = []
        self.trio_hbl_paths: list[str] = []
        self.trio_mbl_path: str = ""
        
        # Shakti UI Variables
        self.var_zoho_file = tk.StringVar()
        self.lbl_zoho_status: tk.Label = None # type: ignore

        self.var_user = tk.StringVar(value="Ashish (CSN)")
        self.var_month = tk.StringVar()
        self.var_mode = tk.StringVar(value="Sea (FCL)")
        self.var_branch = tk.StringVar(value="MUMBAI")

        # UI Widget Attributes
        self.tree: ttk.Treeview = None  # type: ignore
        self.lbl_file_status: tk.Label = None  # type: ignore
        self.cb_user: ttk.Combobox = None  # type: ignore
        self.cal_pre_alert: BetterDateEntry = None  # type: ignore
        self.cal_vessel_eta: BetterDateEntry = None  # type: ignore
        self.cb_month: ttk.Combobox = None  # type: ignore
        self.cb_mode: ttk.Combobox = None  # type: ignore
        self.entry_branch: ttk.Entry = None  # type: ignore
        self.btn_review: tk.Button = None  # type: ignore
        self.btn_push: tk.Button = None  # type: ignore
        self.btn_convert_zoho: tk.Button = None  # type: ignore
        self.footer_btn_frame: tk.Frame = None  # type: ignore

        self._load_logo()
        self._build_header()
        
        # Main Tab Container
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Tab 1: DSR Extractor (Existing logic)
        self.tab_extractor = tk.Frame(self.notebook, bg=WHITE)
        self.notebook.add(self.tab_extractor, text="  Invoice, MBL, HBL Extractor  ")
        self._build_extractor_tab()

        # Tab 2: Shakti Converter (NEW)
        self.tab_zoho = tk.Frame(self.notebook, bg=WHITE)
        self.notebook.add(self.tab_zoho, text="  Shakti export file to DSR's  ")
        self._build_zoho_tab()

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self._build_footer()

    def _load_logo(self) -> None:
        self.logo_img = None
        if LOGO_PATH.exists():
            try:
                img = Image.open(LOGO_PATH)
                img = img.resize((150, 22), Image.LANCZOS)
                self.logo_img = ImageTk.PhotoImage(img)
            except Exception:
                pass

    def _build_header(self) -> None:
        header = tk.Frame(self.root, bg=WHITE, height=80)
        header.pack(fill="x", side="top", pady=10)
        
        # Logo on left
        if self.logo_img:
            lbl_logo = tk.Label(header, image=self.logo_img, bg=WHITE)
            lbl_logo.pack(side="left", padx=20)
        
        # Titles strictly centered
        title_frame = tk.Frame(header, bg=WHITE)
        title_frame.pack(side="top", expand=True)

        tk.Label(
            title_frame, text="Skoda DSR Generator",
            font=("Segoe UI", 20, "bold"), fg=BTN_BLUE, bg=WHITE
        ).pack()

        tk.Label(
            title_frame, text="Extract container-wise data from Bills of Lading and Invoices",
            font=("Segoe UI", 11), fg="#6c757d", bg=WHITE
        ).pack(pady=(2, 0))

    def _build_extractor_tab(self) -> None:
        self._build_file_selection(self.tab_extractor)
        self._build_manual_settings(self.tab_extractor)
        self._build_action_buttons(self.tab_extractor)
        self._build_preview(self.tab_extractor)

    def _build_action_buttons(self, parent_frame: tk.Frame) -> None:
        self.action_btn_frame = tk.Frame(parent_frame, bg=WHITE)
        self.action_btn_frame.pack(fill="x", padx=20, pady=10)

        # Container for the two main action buttons
        btn_container = tk.Frame(self.action_btn_frame, bg=WHITE)
        btn_container.pack(side="right")

        self.btn_review = tk.Button(
            btn_container, text="1. Review & Confirm Data", font=("Segoe UI", 10, "bold"),
            bg="#f39c12", fg=WHITE, activebackground="#e67e22", activeforeground=WHITE,
            width=25, height=2, borderwidth=0, cursor="hand2",
            command=self._on_review
        )
        self.btn_review.pack(side="left", padx=10)

        self.btn_push = tk.Button(
            btn_container, text="2. Push to Shakti & Export", font=("Segoe UI", 10, "bold"),
            bg=BTN_BLUE, fg=WHITE, activebackground="#004494", activeforeground=WHITE,
            width=25, height=2, borderwidth=0, cursor="arrow",
            command=self._on_push_and_export, state="disabled"
        )
        self.btn_push.pack(side="left")

    def _build_file_selection(self, parent_frame: tk.Frame) -> None:
        file_selection_frame = ttk.LabelFrame(parent_frame, text="File Selection", padding=(10, 8))
        file_selection_frame.pack(fill="x", padx=20, pady=5)

        # Simplified Selection: Only Trio Match
        sel_box = tk.Frame(file_selection_frame, bg="#F8F9FA", pady=15, padx=15)
        sel_box.pack(fill="x", pady=5)
        
        row = tk.Frame(sel_box, bg="#F8F9FA")
        row.pack(fill="x")
        
        # Space on left filled by Clear button
        tk.Button(row, text="Clear Inputs", width=12, command=self._on_clear_list, relief="groove").pack(side="left", padx=(0, 20))

        # Invoice
        tk.Button(row, text="Select Invoice", width=12, command=lambda: self._select_trio_file("inv")).pack(side="left", padx=2)
        tk.Entry(row, textvariable=self.var_trio_inv, width=30, font=("Segoe UI", 9)).pack(side="left", padx=2)
        
        # MBL
        tk.Button(row, text="Select MBL", width=12, command=lambda: self._select_trio_file("mbl")).pack(side="left", padx=10)
        tk.Entry(row, textvariable=self.var_trio_mbl, width=28, font=("Segoe UI", 9)).pack(side="left", padx=2)
        
        # HBL
        tk.Button(row, text="Select HBL", width=12, command=lambda: self._select_trio_file("hbl")).pack(side="left", padx=10)
        tk.Entry(row, textvariable=self.var_trio_hbl, width=28, font=("Segoe UI", 9)).pack(side="left", padx=2)
        
        tk.Button(row, text="Extract", bg="#28A745", fg=WHITE, font=("Segoe UI", 9, "bold"), width=15, height=1, command=self._on_process_trio).pack(side="right", padx=(10, 0))
        
        # Action Status below
        self.lbl_file_status = tk.Label(sel_box, text="Ready", font=("Segoe UI", 8), fg="#6c757d", bg="#F8F9FA")
        self.lbl_file_status.pack(side="left", pady=(5,0))

    def _build_manual_settings(self, parent_frame: tk.Frame) -> None:
        frame = ttk.LabelFrame(parent_frame, text="Manual / Shakti Fields", padding=(10, 8))
        frame.pack(fill="x", padx=20, pady=5)
        
        inner = tk.Frame(frame, bg=WHITE)
        inner.pack(fill="x")
        
        # User
        tk.Label(inner, text="User:", font=("Segoe UI", 9, "bold"), bg=WHITE).grid(row=0, column=0, sticky="w", padx=5)
        self.cb_user = ttk.Combobox(inner, textvariable=self.var_user, values=["Ashish (CSN)", "Ranjit (PUNE)", "CLC / After sales"], width=27)
        self.cb_user.grid(row=0, column=1, sticky="w", padx=5)
        
        # Month Dropdown
        tk.Label(inner, text="Month:", font=("Segoe UI", 9, "bold"), bg=WHITE).grid(row=0, column=2, sticky="w", padx=(20, 5))
        self.cb_month = ttk.Combobox(inner, textvariable=self.var_month, values=list(MONTH_MAP.values()), width=10, state="readonly")
        self.cb_month.grid(row=0, column=3, sticky="w", padx=5)

        # Pre-alert Date (custom pop-up)
        tk.Label(inner, text="Pre-alert Receive Date:", font=("Segoe UI", 9, "bold"), bg=WHITE).grid(row=0, column=4, sticky="w", padx=(20, 5))
        self.cal_pre_alert = BetterDateEntry(inner, width=15, bg=WHITE)
        self.cal_pre_alert.grid(row=0, column=5, sticky="w", padx=5)
        
        # Vessel ETA (custom pop-up)
        tk.Label(inner, text="Vessel ETA:", font=("Segoe UI", 9, "bold"), bg=WHITE).grid(row=0, column=6, sticky="w", padx=(20, 5))
        self.cal_vessel_eta = BetterDateEntry(inner, width=15, bg=WHITE)
        self.cal_vessel_eta.grid(row=0, column=7, sticky="w", padx=5)
        
        # Mode Dropdown
        tk.Label(inner, text="Mode:", font=("Segoe UI", 9, "bold"), bg=WHITE).grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.cb_mode = ttk.Combobox(inner, textvariable=self.var_mode, values=["Air", "Sea (FCL)", "Sea (LCL)", "Sea (BB)"], width=15, state="readonly")
        self.cb_mode.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        # Branch (Locked to MUMBAI)
        tk.Label(inner, text="Branch:", font=("Segoe UI", 9, "bold"), bg=WHITE).grid(row=1, column=2, sticky="w", padx=(20, 5), pady=5)
        self.entry_branch = ttk.Entry(inner, textvariable=self.var_branch, width=15, state="readonly")
        self.entry_branch.grid(row=1, column=3, sticky="w", padx=5, pady=5)


    def _build_preview(self, parent_frame: tk.Frame) -> None:
        frame = ttk.LabelFrame(parent_frame, text="Data Preview / Processing Queue", padding=(10, 8))
        frame.pack(fill="both", expand=True, padx=20, pady=5)

        columns = ("Directory", "Files", "Status", "Parsed Container(s)", "Invoice Nos", "BL No", "Action")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings")
        self.tree.tag_configure("duplicate", background="#FFF3CD") # Light yellow for duplicates
        
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)

        col_widths = {
            "Directory": 200, "Files": 80, "Status": 100, 
            "Parsed Container(s)": 150, "Invoice Nos": 200, "BL No": 120,
            "Action": 80
        }
        for col in columns:
            self.tree.heading(col, text=col)
            anchor = "center" if col == "Action" else "w"
            self.tree.column(col, width=col_widths.get(col, 100), anchor=anchor)

        self.tree.bind("<Button-1>", self._on_tree_click)

    def _build_footer(self) -> None:
        footer = tk.Frame(self.root, bg=WHITE, height=40)
        footer.pack(fill="x", side="bottom", padx=20, pady=5)
        
        tk.Label(footer, text="© Nagarkot Forwarders Pvt Ltd", font=("Segoe UI", 8), fg="#6c757d", bg=WHITE).pack(side="left")

    def _on_tab_changed(self, event) -> None:
        """Tab change logic can handle visibility if needed, but actions are embedded in Tab 1."""
        pass

    def _build_zoho_tab(self) -> None:
        """Builds the UI for converting Shakti Exported Excel files."""
        container = tk.Frame(self.tab_zoho, bg=WHITE)
        container.pack(fill="both", expand=True, padx=40, pady=40)

        card = tk.Frame(container, bg=BG_LIGHT, bd=1, relief="solid")
        card.pack(pady=20, padx=20, ipadx=20, ipady=20)

        tk.Label(
            card, text="Shakti export file to DSR Converter", 
            font=("Segoe UI", 16, "bold"), fg=BTN_BLUE, bg=BG_LIGHT
        ).pack(pady=(0, 10))

        tk.Label(
            card, text="Select the Excel file exported from Shakti to generate user-wise DSRs.",
            font=("Segoe UI", 10), bg=BG_LIGHT, fg="#6c757d"
        ).pack(pady=(0, 20))

        sel_row = tk.Frame(card, bg=BG_LIGHT)
        sel_row.pack(fill="x", pady=10)

        tk.Button(
            sel_row, text="Browse Shakti Excel", font=("Segoe UI", 10),
            width=20, command=self._on_select_zoho_file, cursor="hand2"
        ).pack(side="left", padx=5)

        tk.Entry(
            sel_row, textvariable=self.var_zoho_file, font=("Segoe UI", 10), 
            width=60, state="readonly"
        ).pack(side="left", padx=5)

        action_row = tk.Frame(card, bg=BG_LIGHT)
        action_row.pack(fill="x", pady=30)

        self.btn_convert_zoho = tk.Button(
            action_row, text="CONVERT & GENERATE DSR's", font=("Segoe UI", 12, "bold"),
            bg="#28A745", fg=WHITE, width=35, height=2, cursor="hand2",
            command=self._on_convert_zoho
        )
        self.btn_convert_zoho.pack(anchor="center")

        self.lbl_zoho_status = tk.Label(
            card, text="Status: Ready", font=("Segoe UI", 9, "italic"),
            bg=BG_LIGHT, fg="#28A745"
        )
        self.lbl_zoho_status.pack(pady=10)

        # Instructions
        info_frame = tk.Frame(container, bg=WHITE)
        info_frame.pack(fill="x", pady=20)
        
        info_text = (
            "How it works:\n"
            "1. Upload the Shakti export file to DSR Converter export file.\n"
            "2. The tool splits records by User: Ashish (CSn.xlsx), Ranjit (pune.xlsx), CLC (CLC dsr.xlsx).\n"
            "3. Each Excel will have two sheets: 'Live shipments' and 'Cleared shipments'.\n"
            "4. Logic: Records with a 'Dispatch Date' go to 'Cleared', others to 'Live'."
        )
        tk.Label(
            info_frame, text=info_text, font=("Segoe UI", 9),
            justify="left", bg=WHITE, fg="#6c757d"
        ).pack(anchor="w")

    def _on_select_zoho_file(self) -> None:
        f = filedialog.askopenfilename(
            title="Select Shakti Export Excel",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if f:
            try:
                # Quick validation of the export format
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                sheet = wb.active
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                # Normalize headers: strip and convert to string
                headers = [str(cell).strip() if cell else "" for cell in first_row]
                
                if "Container Details - Skoda Container No" not in headers:
                    messagebox.showerror(
                        "Invalid Export Format", 
                        "please export seperate collumn container details excel from shakti"
                    )
                    self.lbl_zoho_status.config(text="Status: Error - Invalid Format", fg="red")
                    self.var_zoho_file.set("")
                    return
                wb.close()
            except Exception as e:
                messagebox.showerror("Error", f"Could not validate Excel file: {e}")
                return

            self.var_zoho_file.set(f)
            self.lbl_zoho_status.config(text=f"Selected: {Path(f).name}", fg=BTN_BLUE)

    def _on_convert_zoho(self) -> None:
        file_path = self.var_zoho_file.get()
        if not file_path:
            messagebox.showwarning("File Missing", "Please select a Shakti export file first.")
            return

        try:
            self.lbl_zoho_status.config(text="Processing... Please wait", fg="#f39c12")
            self.root.update_idletasks()

            input_path = Path(file_path)
            wb_in = openpyxl.load_workbook(input_path, data_only=True)
            sheet = wb_in.active
            
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            
            # Map headers to indices
            h_map = {h: i for i, h in enumerate(headers) if h}
            
            # ─── Validation: Check for separate Container columns ─────────────────
            if "Container Details - Skoda Container No" not in h_map:
                messagebox.showerror(
                    "Invalid Export Format", 
                    "please export seperate collumn container details excel from shakti"
                )
                self.lbl_zoho_status.config(text="Status: Error - Invalid Format", fg="red")
                return
            # ──────────────────────────────────────────────────────────────────────
            
            # Define Shakti to Master mapping based on analysis
            zoho_map = {
                "User": "User",
                "Pre-alert Receive date": "Pre-alert Receive date",
                "Month": "Month",
                "FF/ Shipping Line": "FF/ Shipping Line",
                "Port of Loading": "Port of Loading",
                "Vessel Name": "Vessel Name",
                "BL Date": "BL Date",
                "Vessel ETA": "Vessel ETA",
                "BL No.": "BL No.",
                "CHA Job No.": "CHA Job No.",
                "Current Status": "Current Status",
                "Supplier Name": "Supplier Name",
                "INCO": "INCO",
                "CFS Name": "CFS Name",
                "IGM No.": "IGM No.",
                "IGM No. Date": "IGM No. Date",
                "IGM Inward Date": "IGM Inward Date",
                "B/E No": "B/E No",
                "B/E Date": "B/E Date",
                "AO Ass": "AO Ass",
                "AC Assess": "AC Assess",
                "RMS/ Examine": "RMS/ Examine",
                "Duty Request recd from CHA": "Duty Request recd from CHA",
                "Duty paid date": "Duty Paid date",
                "Assessable Value": "Assessable Value",
                "Debit Duty (RODTEP)": "Debit Duty (RODTEP)",
                "Total Duty": "Total Duty",
                "DUTY%": "DUTY%",
                "Stamp Duty": "STAMP DUTY",
                "Interest (IfAny)": "Interest (IfAny)",
                "Penalty (Ifany)": "Penalty (Ifany)",
                "Reason for Interest / Penalty": "Reason for Interest/Penalty",
                "OOC Date": "OOC Date",
                "Remarks (Daywise Cronology)": "Remarks (Daywise Cronology)",
                "Clearance TAT": "Clearnace TAT",
                "Reason for Clearance TAT Delay": "Reason for Clearance TAT delay",
                "Detention/Demurrage (IfAny)": "Detention/Demurrage (IfAny)",
                "Total BCD Value": "Total BCD Value",
                "Total SWS Value": "Total SWS Value",
                "Total IGST Value": "Total IGST Value",
                "STAMP DUTY PAID DT": "STAMP DUTY PAID DT",
                "Under Protest": "Under Protest",
                "BOE filing TAT": "BOE filing TAT",
                "Reason for BOE filing TAT Delay": "Reason for BOE filing TAT delay",
                "Container arrival date in CFS": "Conatainer arrival date in CFS",
                "OOC COPY RECD YES/NO": "OOC COPY RECD YES/NO",
                "SIMS Registration date": "SIMS Registration date",
                "Remarks": "Remarks",
                "Container Details - Skoda Container No": "Container No.",
                "Container Details - Container Size": "Size (20'40' LCL)",
                "Container Details - Container Type": "Container Type (HQ,DV,SD)",
                "Container Details - Gross Wt": "GrossWt",
                "Container Details - No of Pkgs": "No.of Pkg.",
                "Container Details - Skoda Invoice No": "Invoice No.",
                "Container Details - Skoda Dispatch Date": "Dispatch date to plant/WH",
                "Container Details - Skoda Transporter": "Transporter",
                "Container Details - Skoda E-Waybill No": "E-Waybill No."
            }

            # Master column names to their indices
            master_h_to_idx = {h: i for i, h in enumerate(DSR_HEADERS)}
            
            # User to filename mapping
            curr_date = datetime.now().strftime("%d-%m-%y")
            user_files = {
                "Ashish (CSN)": f"{curr_date} - CSN.xlsx",
                "Ranjit (PUNE)": f"{curr_date} - PUNE.xlsx",
                "CLC / After sales": f"{curr_date} - CLC.xlsx"
            }
            
            # Subsets for each user
            user_data = {u: [] for u in user_files}
            
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            for r_vals in rows:
                if not any(r_vals): continue
                
                # Extract user
                u_val = str(r_vals[h_map["User"]]).strip() if "User" in h_map else ""
                if u_val not in user_data:
                    # Handle unknown users or mismatches if any
                    continue
                
                # Date columns in Master DSR that need timestamp stripping
                date_master_headers = {
                    "Pre-alert Receive date", "BL Date", "Vessel ETA",
                    "IGM No. Date", "IGM Inward Date", "B/E Date",
                    "Duty Request recd from CHA", "Duty Paid date",
                    "OOC Date", "Dispatch date to plant/WH",
                    "STAMP DUTY PAID DT", "Conatainer arrival date in CFS",
                    "SIMS Registration date",
                }

                # Build master row
                m_row = [""] * len(DSR_HEADERS)
                for z_h, m_h in zoho_map.items():
                    if z_h in h_map:
                        val = r_vals[h_map[z_h]]
                        # Only clean date values for date columns
                        if m_h in date_master_headers:
                            val = self._clean_date(val)
                        if m_h in master_h_to_idx:
                            m_row[master_h_to_idx[m_h]] = val
                
                # Special cases: CHA JOB NO index 48 (duplicate in master)
                if "CHA Job No." in h_map:
                    val = r_vals[h_map["CHA Job No."]]
                    if "CHA JOB NO" in master_h_to_idx:
                         m_row[master_h_to_idx["CHA JOB NO"]] = val

                user_data[u_val].append(m_row)

            # Generate files
            out_dir = SCRIPT_DIR
            generated = []

            for user, rows in user_data.items():
                if not rows: continue
                
                fname = user_files[user]
                fpath = out_dir / fname
                
                wb_out = openpyxl.Workbook()
                # Create Sheets
                ws_live = wb_out.active
                ws_live.title = "Live shipments"
                ws_cleared = wb_out.create_sheet("Cleared shipments")
                
                # Headers
                ws_live.append(DSR_HEADERS)
                ws_cleared.append(DSR_HEADERS)
                
                # Data Distribution
                dispatch_date_idx = master_h_to_idx.get("Dispatch date to plant/WH", -1)
                
                for r in rows:
                    is_cleared = False
                    if dispatch_date_idx != -1:
                        d_val = r[dispatch_date_idx]
                        if d_val and str(d_val).strip():
                            is_cleared = True
                    
                    if is_cleared:
                        ws_cleared.append(r)
                    else:
                        ws_live.append(r)
                
                self._apply_dsr_styling(ws_live)
                self._apply_dsr_styling(ws_cleared)
                
                wb_out.save(fpath)
                generated.append(fname)

            self.lbl_zoho_status.config(text="Success: Files Generated!", fg="#28A745")
            msg = "The following files were created in the application directory:\n\n" + "\n".join(generated)
            messagebox.showinfo("Conversion Complete", msg)

        except Exception as e:
            logger.exception("Shakti conversion failed")
            self.lbl_zoho_status.config(text="Error occurred", fg="red")
            messagebox.showerror("Error", f"Failed to convert Shakti file:\n{str(e)}")

    def _on_select_pdfs(self) -> None:
        files = filedialog.askopenfilenames(
            title="Select BL & Invoice PDFs",
            filetypes=[("PDF Files", "*.pdf")],
        )
        if files:
            for f in files:
                p = Path(f)
                d = p.parent
                if d not in self.files_by_dir:
                    self.files_by_dir[d] = []
                if p not in self.files_by_dir[d]:
                    self.files_by_dir[d].append(p)
            self._parse_and_refresh()

    def _on_select_folder(self) -> None:
        folder = filedialog.askdirectory(title="Select folder containing shipment PDFs")
        if folder:
            p = Path(folder)
            pdfs = list(p.glob("*.pdf"))
            if pdfs:
                if p not in self.files_by_dir:
                    self.files_by_dir[p] = []
                for f in pdfs:
                    if f not in self.files_by_dir[p]:
                        self.files_by_dir[p].append(f)
                self._parse_and_refresh()
            else:
                messagebox.showinfo("No PDFs", "No PDF files found in the selected folder.")

    def _select_trio_file(self, kind: str) -> None:
        if kind in ["inv", "hbl"]:
            title = "Select Invoice PDF(s)" if kind == "inv" else "Select HBL PDF(s)"
            files = filedialog.askopenfilenames(title=title, filetypes=[("PDF Files", "*.pdf")])
            if files:
                if kind == "inv":
                    self.trio_inv_paths = list(files)
                    self.var_trio_inv.set(f"({len(files)} files) " + ", ".join(Path(f).name for f in files))
                else:
                    self.trio_hbl_paths = list(files)
                    self.var_trio_hbl.set(f"({len(files)} files) " + ", ".join(Path(f).name for f in files))
        else:
            f = filedialog.askopenfilename(title="Select MBL PDF", filetypes=[("PDF Files", "*.pdf")])
            if f:
                self.trio_mbl_path = f
                self.var_trio_mbl.set(Path(f).name)

    def _on_process_trio(self) -> None:
        try:
            inv_paths = [Path(p) for p in self.trio_inv_paths]
            hbl_paths = [Path(p) for p in self.trio_hbl_paths]
            mbl_path = Path(self.trio_mbl_path) if self.trio_mbl_path else None
            
            # inv_str = self.var_trio_inv.get().strip()
            # hbl_str = self.var_trio_hbl.get().strip()
            # mbl_str = self.var_trio_mbl.get().strip()
            
            # # Invoice(s) can be multiple
            # inv_paths = [Path(p.strip()) for p in inv_str.split(",") if p.strip()]
            
            if not inv_paths or (not hbl_paths and not mbl_path):
                messagebox.showwarning("Missing Files", "Minimum one Invoice + one BL (HBL or MBL) required.")
                return

            # 1. Parse BLs
            mbl_recs = parse_bl(mbl_path) if mbl_path else []
            hbl_recs_all = []
            for hp in hbl_paths:
                hbl_recs_all.extend(parse_bl(hp))
            
            # Combine unique containers from all BLs
            base_recs = []
            seen_c = set()
            raw_recs = mbl_recs + hbl_recs_all
            
            # Filter: if any record has a container, ignore empty ones
            has_containers = any(r.container_no.strip() for r in raw_recs)
            
            for r in raw_recs:
                cur_c = r.container_no.upper().strip()
                if has_containers and not cur_c:
                    continue # Ignore fallback records if we have real details
                    
                if cur_c not in seen_c:
                    base_recs.append(r)
                    seen_c.add(cur_c)
            
            if not base_recs:
                 messagebox.showerror("No Data", "Could not extract container details from the provided BLs.")
                 return

            # Auto-detect month from BL date
            first_bl_date = next((r.bl_date for r in raw_recs if r.bl_date), None)
            if first_bl_date:
                try:
                    dt_obj = None
                    for fmt in ("%Y-%m-%d", "%d-%b-%Y"):
                        try:
                            dt_obj = datetime.strptime(first_bl_date, fmt)
                            break
                        except:
                            continue
                    if dt_obj:
                        m_idx = dt_obj.strftime("%m")
                        if m_idx in MONTH_MAP:
                            self.var_month.set(MONTH_MAP[m_idx])
                except:
                    pass

            # Combine them
            
            # Identify MBL and HBL numbers (Formatted)
            mbl_no_raw = mbl_recs[0].bl_no if mbl_recs else ""
            mbl_raw_keep = mbl_recs[0].raw_mbl_no if mbl_recs else ""
            hbl_nos_list = list(dict.fromkeys(r.bl_no for r in hbl_recs_all))
            
            mbl_no = self._format_bl_number(mbl_no_raw)
            hbl_no_combined = "/".join(self._format_bl_number(h) for h in hbl_nos_list)
            
            # Parse ALL Invoices for items and possible supplier
            inv_text = ""
            for inv_p in inv_paths:
                try:
                    doc = fitz.open(str(inv_p))
                    inv_text += "".join(page.get_text().upper() for page in doc)
                    doc.close()
                except Exception as e:
                    logger.warning(f"Could not read invoice {inv_p}: {e}")
            
            # Detect Supplier (from user instructions)
            detected_supplier = ""
            # Mapping logic from HBL if available (User asked: "take the supplier name mapping from hbl")
            if hbl_recs_all and hbl_recs_all[0].supplier_name:
                detected_supplier = hbl_recs_all[0].supplier_name
            elif "PREMIUM SOUND" in inv_text:
                detected_supplier = "PREMIUM SOUND SOLUTIONS SDN BHD"
            elif "AUDI HUNGARIA" in inv_text:
                detected_supplier = "AUDI HUNGARIA ZRT."
            elif "VOLKSWAGEN AG" in inv_text:
                detected_supplier = "VOLKSWAGEN AG"
            elif "AUDI AG" in inv_text:
                detected_supplier = "AUDI AG"
            elif "SKODA AUTO" in inv_text or "CELKOV" in inv_text:
                detected_supplier = "Skoda Auto A.S."
            
            # 2. Map Invoices to Containers (Logic from _parse_and_refresh)
            container_to_invoices = {rec.container_no.upper(): set() for rec in base_recs}
            unmapped_invoices = set()
            global_inv_nos = [] # All detected invoice numbers

            for inv_p in inv_paths:
                # Determine Invoice Number (Priority: Content -> Filename -> Stem)
                detected_inv_nos = []
                inv_raw_text = ""
                try:
                    doc = fitz.open(str(inv_p))
                    for page in doc:
                        text_chunk = page.get_text().upper()
                        inv_raw_text += text_chunk + "\n"
                        found = re.findall(r"INVOICE\s*(?:NO|NUMBER|#)?\.?\s*[:\-]?\s*(\d{8,10})", text_chunk)
                        if found: detected_inv_nos.extend(found)
                    doc.close()
                except: pass

                if not detected_inv_nos:
                    detected_inv_nos = re.findall(r"\b(\d{8,10})\b", inv_p.stem)
                
                inv_no = "/".join(dict.fromkeys(detected_inv_nos)) if detected_inv_nos else re.split(r"[-.]", inv_p.stem)[0]
                global_inv_nos.append(inv_no)

                # Find containers in this specific invoice
                found_containers = set(re.findall(r"\b([A-Z]{4}\d{7})\b", inv_raw_text))
                mapped = False
                for c_no in found_containers:
                    if c_no in container_to_invoices:
                        container_to_invoices[c_no].add(inv_no)
                        mapped = True
                
                if not mapped:
                    unmapped_invoices.add(inv_no)

            # 3. Finalize Records
            combined_records = []
            for r in base_recs:
                r.hbl_no = hbl_no_combined
                r.mbl_no = mbl_no
                r.raw_mbl_no = mbl_raw_keep
                # Combined BL Pattern: MBL/HBL as requested
                r.bl_no = f"{mbl_no}/{hbl_no_combined}" if mbl_no and hbl_no_combined else (mbl_no or hbl_no_combined)
                
                cno = r.container_no.upper()
                mapped_set = container_to_invoices.get(cno, set())
                
                if mapped_set:
                    # Invoices found FOR THIS container + any global/unmapped ones
                    final_invs = sorted(list(mapped_set | unmapped_invoices))
                    r.invoice_nos = "/".join(final_invs)
                else:
                    # Fallback: if no specific mapping found, use all (or unmapped)
                    r.invoice_nos = "/".join(sorted(list(unmapped_invoices))) if unmapped_invoices else "/".join(dict.fromkeys(global_inv_nos))

                if detected_supplier:
                    r.supplier_name = detected_supplier
                combined_records.append(r)

            if not combined_records:
                messagebox.showerror("No Data", "Could not extract container details from the provided BLs.")
                return

            self.parsed_records.extend(combined_records)
            
            # Add to tree
            inv_names = ", ".join(p.name for p in inv_paths)
            display_name = inv_names if len(inv_names) < 40 else f"{inv_names[:37]}..."
            
            total_files = len(inv_paths) + len(hbl_paths) + (1 if mbl_path else 0)
            containers_str = ", ".join(r.container_no for r in base_recs)
            display_bl = base_recs[0].bl_no if base_recs else "None"
            all_invs_display = "/".join(dict.fromkeys(global_inv_nos))
            
            # Get directory name from the first file path
            dir_name = "Unknown"
            if inv_paths:
                dir_name = inv_paths[0].parent.name
            elif hbl_paths:
                dir_name = hbl_paths[0].parent.name
            elif mbl_path:
                dir_name = Path(mbl_path).parent.name

            item_iid = self.tree.insert("", "end", values=(f"{dir_name} ({total_files} Files)", f"{total_files} PDFs", "Completed", containers_str, all_invs_display, display_bl, "✖ Remove"))
            self.tree.item(item_iid, tags=("TRIO_SOURCE",))
            
            # Clear trio fields
            self.trio_inv_paths.clear()
            self.trio_hbl_paths.clear()
            self.trio_mbl_path = ""
            self.var_trio_inv.set("")
            self.var_trio_hbl.set("")
            self.var_trio_mbl.set("")
            messagebox.showinfo("Success", f"Extracted {len(combined_records)} container(s).")

        except Exception as e:
            logger.exception("Trio extraction failed")
            messagebox.showerror("Extraction Error", str(e))

    def _on_clear_list(self) -> None:
        self.files_by_dir.clear()
        self.parsed_records.clear()
        self.confirmed_records.clear()
        self.tree.delete(*self.tree.get_children())
        self.lbl_file_status.config(text="No files selected")
        self.btn_push.config(state="disabled", cursor="arrow")


    def _parse_and_refresh(self) -> None:
        self.tree.delete(*self.tree.get_children())
        self.parsed_records.clear()
        
        total_files = sum(len(flist) for flist in self.files_by_dir.values())
        self.lbl_file_status.config(text=f"{total_files} file(s) across {len(self.files_by_dir)} folder(s)")

        # To detect duplicates
        seen_combos = {} # (containers_str, bl_no) -> list of iids

        for directory, files in self.files_by_dir.items():
            # Identify BL vs Invoice
            bl_files = []
            invoice_pdf_files = []
            
            for f in files:
                stem_upper = f.stem.upper()
                if (stem_upper.startswith("MAEU") or stem_upper.startswith("HLCU") 
                    or stem_upper == "BL" or stem_upper.startswith("MEAU")
                    or stem_upper.startswith("SWB")):
                    bl_files.append(f)
                else:
                    invoice_pdf_files.append(f)
            
            # Parse BLs
            dir_records = []
            for bl_file in bl_files:
                try:
                    recs = parse_bl(bl_file)
                    dir_records.extend(recs)
                except Exception as e:
                    logger.error(f"Error parsing {bl_file.name}: {e}")
                    status = f"Error parsing BL"
            
            container_to_invoices = {rec.container_no.upper(): set() for rec in dir_records}
            container_to_supplier = {}
            unmapped_invoices = set()
            unmapped_supplier = None  # Supplier detected from invoices that have no container numbers
            all_invoice_stems = []
            
            for inv_f in invoice_pdf_files:
                # Improved Invoice extraction: handle 8-10 digit numbers
                matches = re.findall(r"\b(\d{8,10})\b", inv_f.stem)
                if matches:
                    inv_no = "/".join(dict.fromkeys(matches))
                else:
                    # Fallback: take the first part before hyphen/dot
                    inv_no = re.split(r"[-.]", inv_f.stem)[0]
                
                all_invoice_stems.append(inv_no)
                
                try:
                    doc = fitz.open(str(inv_f))
                    text_all = "".join(page.get_text().upper() for page in doc)
                    doc.close()
                    
                    found_containers = set(re.findall(r"\b([A-Z]{4}\d{7})\b", text_all))
                    
                    # Detect Supplier from Invoice Content (Priority Check)
                    detected_supplier = None
                    if "PREMIUM SOUND" in text_all:
                        detected_supplier = "PREMIUM SOUND SOLUTIONS SDN BHD"
                    elif "AUDI HUNGARIA" in text_all:
                        detected_supplier = "AUDI HUNGARIA ZRT."
                    elif "VOLKSWAGEN AG" in text_all:
                        detected_supplier = "VOLKSWAGEN AG"
                    elif "AUDI AG" in text_all:
                        detected_supplier = "AUDI AG"
                    elif "SKODA AUTO" in text_all or "CELKOV" in text_all or "IN SEA" in inv_f.stem.upper():
                        detected_supplier = "Skoda Auto A.S."

                    mapped = False
                    for c_no in found_containers:
                        if c_no in container_to_invoices:
                            container_to_invoices[c_no].add(inv_no)
                            if detected_supplier:
                                container_to_supplier[c_no] = detected_supplier
                            mapped = True
                    
                    if not mapped:
                        # Invoice has no matching container (e.g. Skoda AS invoices)
                        unmapped_invoices.add(inv_no)
                        if detected_supplier:
                            unmapped_supplier = detected_supplier
                        
                except Exception as e:
                    logger.error(f"Error parsing invoice {inv_f.name}: {e}")
                    unmapped_invoices.add(inv_no)

            all_invoice_stems = list(dict.fromkeys(all_invoice_stems))
            invoices_str = "/".join(all_invoice_stems)

            status = "Parsed"
            containers_str = "-"
            bl_no = "-"

            if not bl_files:
                status = "Error: No BL found"
                item_iid = self.tree.insert("", "end", values=(directory.name, f"{len(files)} files", status, containers_str, invoices_str, bl_no, "✖ Remove"))
                self.tree.item(item_iid, tags=(str(directory),)) # Save path in tags
                continue
                
            # Apply invoice mappings to the parsed base records
            if invoice_pdf_files:
                for rec in dir_records:
                    cno = rec.container_no.upper()
                    mapped_set = container_to_invoices.get(cno, set())
                    
                    # Update Supplier from container-specific invoice mapping first
                    if cno in container_to_supplier:
                        rec.supplier_name = container_to_supplier[cno]
                    elif unmapped_supplier:
                        # Fallback: Use supplier from unmapped invoices (e.g. Skoda AS with no containers)
                        rec.supplier_name = unmapped_supplier
                    
                    if mapped_set:
                        # Exclusively use invoices mapped specifically to this container
                        all_for_container = sorted(list(mapped_set))
                        # Also append unmapped invoices to each container
                        if unmapped_invoices:
                            all_for_container = sorted(list(mapped_set | unmapped_invoices))
                        rec.invoice_nos = "/".join(all_for_container)
                    else:
                        # No container-specific mapping: assign all unmapped or all invoices
                        if unmapped_invoices:
                            rec.invoice_nos = "/".join(sorted(list(unmapped_invoices)))
                        else:
                            rec.invoice_nos = invoices_str

            self.parsed_records.extend(dir_records)

            if dir_records and not self.var_month.get():
                m_date = dir_records[0].bl_date
                if m_date and len(m_date) >= 7:
                    auto_month = MONTH_MAP.get(m_date[5:7], "")
                    if auto_month:
                        self.var_month.set(auto_month)

            if dir_records:
                containers_str = ", ".join(r.container_no for r in dir_records)
                bl_nos = list(dict.fromkeys(self._format_bl_number(r.bl_no) for r in dir_records))
                bl_no = ", ".join(bl_nos)

            item_iid = self.tree.insert("", "end", values=(directory.name, f"{len(files)} PDFs", status, containers_str, invoices_str, bl_no, "✖ Remove"))
            self.tree.item(item_iid, tags=(str(directory),)) # Hide exact path in tags

            # Duplicate Highlighting Logic
            combo_key = f"{containers_str}|{bl_no}"
            if combo_key in seen_combos:
                # This is a duplicate. Mark both the previous ones and this one.
                for prev_iid in seen_combos[combo_key]:
                    existing_tags = list(self.tree.item(prev_iid, "tags"))
                    if "duplicate" not in existing_tags:
                        self.tree.item(prev_iid, tags=tuple(existing_tags + ["duplicate"]))
                
                existing_tags = list(self.tree.item(item_iid, "tags"))
                self.tree.item(item_iid, tags=tuple(existing_tags + ["duplicate"]))
                seen_combos[combo_key].append(item_iid)
            else:
                seen_combos[combo_key] = [item_iid]

    def _on_tree_click(self, event) -> None:
        """Handles removal of a row if the 'Action' column is clicked."""
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
            
        column = self.tree.identify_column(event.x)
        if column == "#7": # Action column
            item_iid = self.tree.identify_row(event.y)
            if not item_iid:
                return
            
            item_tags = self.tree.item(item_iid, "tags") or []
            
            # Identify if it's a Trio entry (tag "TRIO_SOURCE") or a standard one (tag is directory string)
            if "TRIO_SOURCE" in item_tags:
                if messagebox.askyesno("Remove", "Remove this Trio extraction row?"):
                    # For Trio, we need to find and remove the specific records from parsed_records
                    # This is trickier since TRIO doesn't map to a single dir.
                    # Simple approach: remove by row contents match (container_no + bl_no)
                    values = self.tree.item(item_iid, "values")
                    target_containers = values[3].split(", ")
                    target_bl = values[5]
                    
                    self.parsed_records = [r for r in self.parsed_records if not (r.container_no in target_containers and r.bl_no == target_bl)]
                    self.tree.delete(item_iid)
            else:
                # Standard folder/file removal
                if not item_tags:
                    return
                dir_str = item_tags[0]
                dir_path = Path(dir_str)
                if dir_path in self.files_by_dir:
                    if messagebox.askyesno("Remove", f"Remove selection for '{dir_path.name}'?"):
                        # Remove files from mapping
                        del self.files_by_dir[dir_path]
                        # Re-parse everything else to keep parity
                        self._parse_and_refresh()

    def _on_review(self) -> None:
        global_user = self.var_user.get().strip()
        global_month = self.var_month.get().strip()
        global_pre = self.cal_pre_alert.get().strip()
        global_eta = self.cal_vessel_eta.get().strip()
        global_bl_mode = self.var_mode.get()
        
        if not global_pre or not global_eta:
            messagebox.showwarning("Input Required", "Please select 'Pre-alert Receive Date' and 'Vessel ETA' first.")
            return

        defaults = {
            "user": global_user,
            "user_month": global_month,
            "pre_alert_date": global_pre,
            "vessel_eta": global_eta,
            "bl_mode": global_bl_mode
        }
        
        for r in self.parsed_records:
            r.user = global_user
            r.user_month = global_month
            r.pre_alert_date = global_pre
            r.vessel_eta = global_eta
            r.bl_mode = global_bl_mode

        # Pop up the Data Review modal. 
        DataPreviewWindow(self.root, self.parsed_records, self._on_confirmation_complete, defaults=defaults)

    def _on_confirmation_complete(self) -> None:
        """Called after user finishes editing/confirming in the Review window."""
        self.confirmed_records = list(self.parsed_records)
        self.btn_push.config(state="normal", cursor="hand2")
        
        # Refresh the main GUI treeview to show any manual edits made
        self._refresh_treeview()
        
        messagebox.showinfo("Ready", "Data confirmed! You can now click '2. Push to Shakti & Export'.")

    def _refresh_treeview(self) -> None:
        """Updates the main GUI treeview to reflect manual edits made in the Review window."""
        # For simplicity, we just aggregate all currently parsed BL Nos 
        # and update the BL No column in all existing rows.
        all_bls = list(dict.fromkeys(self._format_bl_number(r.bl_no) for r in self.parsed_records if r.bl_no))
        bl_str = ", ".join(all_bls)
        
        for item in self.tree.get_children():
            vals = list(self.tree.item(item, "values"))
            if len(vals) >= 6:
                vals[5] = bl_str
                self.tree.item(item, values=vals)

    def _get_existing_invoices(self) -> set[tuple[str, str]]:
        """Reads the local Master DSR and returns a set of (Container No, Invoice No) tuples."""
        existing_keys = set()
        if not MASTER_FILE_PATH.exists():
            return existing_keys
            
        try:
            wb = openpyxl.load_workbook(MASTER_FILE_PATH, read_only=True)
            # Find common sheet names
            sheet_name = next((n for n in ["Live shipments", "DSR"] if n in wb.sheetnames), wb.sheetnames[0])
            ws = wb[sheet_name]
            
            # DSR_HEADERS: J=9 (Container No), P=15 (Invoice No) (0-indexed)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > 15:
                    c_no = str(row[9]).strip().upper() if row[9] else ""
                    inv_no_str = str(row[15]).strip().upper() if row[15] else ""
                    if c_no and inv_no_str:
                        # Split multiple invoices to check individually
                        for single_inv in inv_no_str.split("/"):
                            existing_keys.add((c_no, single_inv.strip()))
            wb.close()
        except Exception as e:
            logger.warning(f"Could not read master file for duplicate check: {e}")
        return existing_keys

    def _on_push_and_export(self) -> None:
        if not self.confirmed_records:
            messagebox.showwarning("Error", "Please review and confirm data first.")
            return

        # 1. Duplicate Check using local Master file
        existing_keys = self._get_existing_invoices()
        records_to_process = []
        duplicate_count = 0
        
        for rec in self.confirmed_records:
            c_no = (rec.container_no or "").strip().upper()
            inv_str = (rec.invoice_nos or "").strip().upper()
            
            # Check if this container+invoice combo already exists
            is_new = False
            for single_inv in inv_str.split("/"):
                inv_key = single_inv.strip()
                if not inv_key or (c_no, inv_key) not in existing_keys:
                    is_new = True
                    break
            
            if not is_new and inv_str: # If ALL invoices for this container are already there
                duplicate_count += 1
            else:
                records_to_process.append(rec)

        if not records_to_process:
            messagebox.showinfo("Duplicates Found", f"All {duplicate_count} records were already found in the Master DSR.\nNo new data to push.")
            return

        if duplicate_count > 0:
            if not messagebox.askyesno("Duplicates Found", f"{duplicate_count} records appear to be already in the Master DSR.\n\nContinue pushing {len(records_to_process)} records?"):
                return

        # 2. Push to Shakti
        try:
            zoho = ZohoCreatorAPI()
            success, msg = zoho.push_records(records_to_process)
            
            if not success:
                messagebox.showerror("Shakti Push Failed", f"Excel export aborted because Shakti push failed:\n\n{msg}")
                return
                
            shakti_msg = f"Shakti API: {msg}"
            
        except Exception as se:
            messagebox.showerror("Shakti API Error", f"Excel export aborted due to API error:\n\n{se}")
            return

        # 3. If Shakti is successful, Update Master Excel
        try:
            if MASTER_FILE_PATH.exists():
                self._append_to_master(MASTER_FILE_PATH, records_to_process)
                excel_msg = "Data successfully appended to Master DSR!"
            else:
                self._create_new_dsr(MASTER_FILE_PATH, records_to_process)
                excel_msg = "Master DSR created and data saved successfully!"
            
            messagebox.showinfo("Success", f"{excel_msg}\n\n{shakti_msg}")
            
            # Reset workflow
            self.confirmed_records.clear()
            self.btn_push.config(state="disabled", cursor="arrow")

        except Exception as exc:
            logger.exception("Failed to update Master Excel")
            messagebox.showerror("Excel Error", f"An error occurred after Shakti push while updating the Master file:\n{exc}")

    # ── Excel Export Logic ───────────────────────────────────────────────

    def _format_bl_number(self, bl_num_str: str) -> str:
        """Strips 'MAEU' prefix from Maersk BLs, keeps others."""
        if not bl_num_str:
            return ""
        parts = bl_num_str.split('/')
        formatted_parts = []
        for part in parts:
            part = part.strip()
            if part.upper().startswith("MAEU"):
                formatted_parts.append(part[4:]) # Strip MAEU
            else:
                formatted_parts.append(part)
        return "/".join(formatted_parts)

    def _clean_date(self, val):
        """Standardizes dates by stripping timestamps and returning date objects where possible."""
        if not val:
            return None
        if isinstance(val, datetime):
            return val.date()
        if not isinstance(val, str):
            return val
            
        # Strip timestamp if present "YYYY-MM-DD 00:00:00" or "DD-MMM-YYYY 00:00:00"
        date_str = val.split(" ")[0] if " " in val else val
        
        # Try various formats
        for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return date_str

    def _record_to_row(self, rec: ContainerRecord) -> list:
        row = [""] * len(DSR_HEADERS)
        
        row[0] = rec.user
        row[1] = self._clean_date(rec.pre_alert_date)
        row[2] = rec.user_month
        row[3] = rec.shipping_line
        row[4] = rec.port_of_loading
        row[5] = rec.vessel_name

        row[6] = self._clean_date(rec.bl_date)
        row[7] = self._clean_date(rec.vessel_eta)
        row[9] = rec.container_no
        row[10] = rec.container_size
        row[11] = rec.container_type
        
        # BL No Column (N) - Combined Pattern (MBL/HBL) as requested
        mbl_formatted = self._format_bl_number(rec.mbl_no)
        hbl_formatted = self._format_bl_number(rec.hbl_no)

        if mbl_formatted and hbl_formatted:
            row[13] = f"{mbl_formatted}/{hbl_formatted}"
        else:
            row[13] = self._format_bl_number(rec.bl_no) # Fallback to general bl_no, also formatted
            
        row[14] = rec.supplier_name
        row[15] = rec.invoice_nos
        row[16] = rec.inco_terms

        try: row[17] = int(rec.num_packages)
        except: row[17] = rec.num_packages

        try: row[18] = float(rec.gross_weight)
        except: row[18] = rec.gross_weight

        return row

    def _apply_dsr_styling(self, ws: Worksheet) -> None:
        """Applies NAGARKOT styling to the DSR sheet."""
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1B3A5C")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        data_font = Font(name="Segoe UI", size=9)
        data_align = Alignment(vertical="center", wrap_text=False)

        # Style header
        for col_idx, _ in enumerate(DSR_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[1].height = 40
        ws.freeze_panes = "A2"

        # Style data rows
        # Explicit list of date columns (must match DSR_HEADERS exactly)
        date_header_names = {
            "Pre-alert Receive date", "BL Date", "Vessel ETA",
            "IGM No. Date", "IGM Inward Date", "B/E Date",
            "Duty Request recd from CHA", "Duty Paid date",
            "OOC Date", "Dispatch date to plant/WH",
            "STAMP DUTY PAID DT", "Conatainer arrival date in CFS",
            "SIMS Registration date",
        }
        date_col_indices = [
            i for i, h in enumerate(DSR_HEADERS, 1) if h in date_header_names
        ]

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                
                # Apply date format DD-MM-YYYY to identified columns
                if cell.column in date_col_indices:
                    cell.number_format = "DD-MM-YYYY"

        # Auto-width
        for col_idx in range(1, len(DSR_HEADERS) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15

    def _append_to_master(self, path: Path, records: list[ContainerRecord]) -> None:
        wb = openpyxl.load_workbook(path)
        sheet_name = next((n for n in ["Live shipments", "DSR"] if n in wb.sheetnames), wb.sheetnames[0])
        ws = wb[sheet_name]
        
        for rec in records:
            ws.append(self._record_to_row(rec))
        
        self._apply_dsr_styling(ws)
        wb.save(path)
        wb.close()

    def _create_new_dsr(self, save_path: Path, records: list[ContainerRecord]) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Live shipments"
        ws.append(DSR_HEADERS)
        
        for rec in records:
            ws.append(self._record_to_row(rec))
            
        self._apply_dsr_styling(ws)
        # Create second sheet empty as per original code pattern
        cleared = wb.create_sheet("Cleared shipments")
        # Apply header to second sheet too
        cleared.append(DSR_HEADERS)
        self._apply_dsr_styling(cleared)

        wb.save(save_path)
        wb.close()


def main() -> None:
    root = tk.Tk()
    app = DSRGeneratorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
